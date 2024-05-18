# -*- coding: UTF-8 -*-
import os, re, argparse
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from tools import read_file, list_con, get_outfile, save_to_xlsx

brackets = '“”'
default_title_pattern = r"^##(.*)第\s*(\d+)\s*章\s*(.*)$"

def is_title(line, title_pattern):
    return len(re.findall(title_pattern, line)) == 1

def extract_bracketed_text(text, title_pattern):
    left_bracket_index = text.find(brackets[0])
    right_bracket_index = text.find(brackets[1])
    return_list = [text]
    if is_title(text, title_pattern):
        return return_list
    # 检查括号是否成对出现
    if left_bracket_index != -1 and right_bracket_index != -1:
        # 检查括号是否匹配
        if left_bracket_index < right_bracket_index:
            return_list = []
            if left_bracket_index > 0:
                return_list.append(text[:left_bracket_index])
            # 提取括号内的内容
            return_list.append(text[left_bracket_index:right_bracket_index+1])
            if right_bracket_index<len(text)-1:
                return_list.extend(extract_bracketed_text(text[right_bracket_index+1:], title_pattern))
    return return_list

def segment_text(filepath, sheet_name='正文白表', fill_color=0, title_pattern=None, **kwargs):    
    lines = read_file(filepath)
    lines = list(filter(lambda x: len(x)>0, lines))
    lines = [s.strip() for s in lines]
    lines = list_con([extract_bracketed_text(s, title_pattern) for s in lines])
    lines = list(filter(lambda x: len(x)>0, lines))

    if title_pattern is not None:
        pattern = title_pattern
    else:
        pattern = default_title_pattern
    titles = list(filter(lambda x: len(re.findall(title_pattern, x))==1, lines))

    # 构建df
    df = pd.DataFrame(lines, columns=['正文'])
    df['集数'] = np.nan
    df['标题'] = np.nan
    df['角色'] = ''

    # 正文以不是以“开头的，角色设置为旁白
    df.loc[df['正文'].apply(lambda x: not re.match(r'^“', x)), '角色'] = '旁白'
    df.loc[df['正文'].isin(titles), '集数'] = df.loc[df['正文'].isin(titles), '正文'].apply(lambda x: f'第{re.findall(pattern, x)[0][1]}章')
    df.loc[df['正文'].isin(titles), '标题'] = df.loc[df['正文'].isin(titles), '正文'].apply(lambda x: re.findall(pattern, x)[0][2].strip(":："))
    df.fillna(method='ffill', inplace=True)
    df = df.reindex(columns=['集数', '标题', '角色', '正文'])
    df = df[~df['正文'].isin(titles)]
    
    print(df.head())

    outfile = get_outfile(filepath)
    if fill_color == 0:
        return save_to_xlsx(df, outfile, sheet_name, mode='w')
    else:
        with pd.ExcelWriter(outfile, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            fill = PatternFill('solid', fgColor='FFFF00')
            for row in range(len(df)-1):  # 行索引从1开始，因为第一行通常是列标题
                if df.iloc[row]['角色'] != '旁白':  # 减1是因为DataFrame的索引是从0开始的
                    for ceil in ws[row+2]:
                        ceil.fill = fill
                # 显示进度
                # if row%100==0: print(f'进度：{row+1}/{len(df)-1}')
                
        return f'表格已保存至 “{outfile}”，表名为 “{sheet_name}”'

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--filepath', type=str, default=None)
    parser.add_argument('-p', '--pattern', type=str, default=None)
    parser.add_argument('-fc', '--fill_color', type=int, default=0)
    args = parser.parse_args()

    # 输入文件路径
    if args.filepath is None:
        filepath = input("请输入文件路径: ")
    else:
        filepath = args.filepath
    filepath = filepath.strip()
    filepath = filepath.replace('"', '')
    if os.path.exists(filepath):
        print(f'文件路径: {filepath}')
    else:
        print(f'文件路径不存在: {filepath}')
        input("按任意键退出...")
        exit()
    message = segment_text(filepath, fill_color=args.fill_color, title_pattern=args.pattern)
    print(message)
    input("按任意键退出...")
